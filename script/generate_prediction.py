import pandas as pd
import joblib
from tensorflow.keras.models import load_model
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from preprocess import load_data, preprocess_data

def save_predictions(original_df, predicted_df, output_path):
    """Save the original and predicted values to an Excel file in alternating rows."""
    # Create a new DataFrame for the results, including only relevant numeric columns
    results_df = pd.DataFrame(columns=['Label'] + list(predicted_df.columns))  # Include 'Label' column

    # Populate DataFrame with alternating original and predicted values
    for i in range(len(original_df)):
        # Append original values
        original_row = {'Label': 'Original'}
        original_row.update(original_df.iloc[i].to_dict())
        results_df = results_df.append(original_row, ignore_index=True)

        # Append predicted values
        predicted_row = {'Label': 'Predicted'}
        predicted_row.update(predicted_df.iloc[i].to_dict())  # Get predicted values for all features
        results_df = results_df.append(predicted_row, ignore_index=True)

    # Create a new workbook and add the results DataFrame
    wb = Workbook()
    ws = wb.active

    # Write DataFrame to Excel
    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws.append(r)

    # Apply color coding for predicted values
    for i in range(len(original_df)):
        for j, target_column in enumerate(predicted_df.columns):
            original_value = original_df[target_column].iloc[i]
            predicted_value = predicted_df[target_column].iloc[i]
            difference = abs(predicted_value - original_value)

            # Determine color based on the difference
            fill_color = '00FF00' if difference < 0.05 else 'FFFF00' if difference < 0.1 else 'FF0000'  # Green, Yellow, Red

            # Color the predicted value cell
            predicted_cell = ws.cell(row=(i * 2 + 3), column=j + 2)  # Adjusting for the label and header
            predicted_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

    # Save the workbook with the formatting
    wb.save(output_path)
    print(f"Predictions saved successfully at {output_path}.")

def main():
    target_columns = [
        'Vibration Frequency',
        'Vibration Amplitude',
        'Bearing Temperature',
        'Motor Temperature',
        'Belt Load',
        'Torque',
        'Noise Levels',
        'Current and Voltage',
        'Hydraulic Pressure',
        'Belt Thickness',
        'Roller Condition'
    ]
    
    models = {}
    scalers = {}
    
    # Load the data
    data = load_data('data/input_data.xlsx')
    
    # Preprocess the data
    processed_data = preprocess_data(data)

    # Load models and scalers for each target column
    for target_column in target_columns:
        models[target_column] = load_model(f'models/{target_column}_model.h5')
        scalers[target_column] = joblib.load(f'models/{target_column}_scaler.pkl')

    # Prepare DataFrame for predictions
    predictions_df = pd.DataFrame(columns=target_columns)

    # Make predictions for the last 10 rows of the data
    last_n_rows = processed_data.tail(10)

    # Filter out only the numeric columns for predictions
    last_n_rows_numeric = last_n_rows[target_columns]

    for target_column in target_columns:
        X = last_n_rows.drop(columns=[target_column])  # Exclude the target column for prediction
        X_scaled = scalers[target_column].transform(X)  # Scale the features
        predictions = models[target_column].predict(X_scaled)

        # Add predictions to the DataFrame
        predictions_df[target_column] = predictions.flatten()  # Flatten the array and add it to the DataFrame

    # Save predictions to Excel with color coding
    save_predictions(last_n_rows_numeric, predictions_df, 'results/predictions.xlsx')

if __name__ == "__main__":
    main()
